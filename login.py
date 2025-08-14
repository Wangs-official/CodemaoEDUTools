# @author:Wangs_official
try:
    from openpyxl import load_workbook
except ImportError:
    exit("请先 pip3 install openpyxl")
try:
    import requests
except ImportError:
    exit("请先 pip3 install requests")
import json
import os
import time

if os.path.exists("tokens.txt"):
    os.remove("tokens.txt")

cannot_login = 0


def login(username, pwd):
    url = "https://api.codemao.cn/tiger/v3/web/accounts/login"
    header = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36"
    }
    post_text = json.dumps({'pid': '65edCTyg', 'identity': username, 'password': pwd})
    return requests.post(url, data=post_text, headers=header)


xlsx_file_path = input(
    "⚠️每次登录都会删掉所保存的tokens.txt文件，请注意\n请输入Excel表路径（Excel的格式应该为.xlsx，文件内格式请查看README）: ")
wb = load_workbook(xlsx_file_path)
sheet = wb.active
all_account = sheet.max_row
print(f"---------------\n总共{all_account}个账号待登录\n---------------")
for _ in range(all_account):
    try:
        start_time = time.time()
        users = []
        passwords = []
        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=3, values_only=True):
            user, password = row
            users.append(user)
            passwords.append(password)
        wb.close()
        nn = login(users[_], passwords[_])
        token = json.loads(nn.text).get("auth", {}).get("token")
        with open("tokens.txt", "a", encoding="utf-8") as f:
            f.write(token + "\n")
        ut = time.time() - start_time
        speed = int(60 / ut)
        print("\r" + f"第{_ + 1}/{all_account}个登录中... | 账号 {users[_]} | 速度 {speed}个/分", end="")       
    except TypeError as e:
        cannot_login = cannot_login + 1
        nnt = str(nn.text)
        print(f"\n此账号登录失败，返回内容：{nnt}\n---------------")

if cannot_login == 0:
    exit("全部登录成功")
else:
    exit(f"有{cannot_login}个账号未正常登录，其余全部登录成功, 请查看上方日志")
