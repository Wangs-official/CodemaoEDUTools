"""
====================
Author: WangZixu
欢迎使用 CodemaoEDUTools!
https://github.com/Wangs-official/CodemaoEDUTools/
====================
请阅读 doc/cli.md 中的命令行参数，或在你的程序中添加此库
import CodemaoEDUTools
====================
开发者不对您使用本项目造成的风险负责，请自行考虑是否使用，谢谢！
====================
请在开始使用前运行
pip3 install -r requirements.txt
====================
"""

import json
import logging
import os
from concurrent.futures import ThreadPoolExecutor
from typing import Any

import argparse
import coloredlogs
import pandas as pd
import requests
from fake_useragent import UserAgent
from openpyxl import Workbook
from openpyxl import load_workbook

coloredlogs.install(level="INFO", fmt="%(asctime)s - %(funcName)s: %(message)s")

version = "1.2.3"
max_workers = 8
student_names = '["xvbnmklq","asdfghjk","qwertyui","zxcvbnml","poiuytre","lkjhgfds","mnbvcxza","plokmijn","uhbygvtd","crfvtgby","edcrfvtg","qazwsxed","rfvtgbyh","nujmikol","zxasqwde","plmnkoij","bvcdxsza","qwermnbp","asxcvgfr","lpoikmju","yhnujmik","tgbzdxew","rfvgyhuj","edcwsxqa","zaqxswcd","vfrcdews","bgtnhyuj","mkiopluj","nhybtgvr","cdexswza","qwerfdsa","zxcvfdsa","poiuytrw","lkjhgfda","mnbvcxzs","asdfqwer","zxcvqwer","poiulkjh","mnbvcxas","qwertzui","yxcvbnmq","plokmnji","uhbgyvft","crfvtgyn","edcrfvbg","qazwsxrf","rfvtgbyu","nujmiklp","zxasqwed","plmnkoji","bvcdxsaz","qwermnbo","asxcvfgd","lpoikmjn","yhnujmki","tgbzdxec","rfvgyhuk","edcwsxqz","zaqxswce","vfrcdewa","bgtnhyum","mkioplun","nhybtgvf","cdexswzb","qwerfdsz","zxcvfdsz","poiuytrq","lkjhgfdz","mnbvcxzc","asdfqwez","zxcvqwez","poiulkjm","mnbvcxaq","qwertzuy","yxcvbnmr","plokmnjh","uhbgyvfr","crfvtgyb","edcrfvbn","qazwsxre","rfvtgbyi","nujmiklj","zxasqweg","plmnkojh","bvcdxsay","qwermnbu","asxcvfgh","lpoikmjh","yhnujmko","tgbzdxer","rfvgyhun","edcwsxqv","zaqxswec","vfrcdewq","bgtnhyup","mkiopluh","nhybtgvc","cdexswzg","qwerfdsx","zxcvfdsx"]'
report_readtoken_line = 20


"""POST方式调用API"""


def PostAPI(Path: str, PostData: dict, Token: str) -> requests.Response:
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": UserAgent().random,
        "authorization": Token,
    }
    return requests.post(
        url=f"https://api.codemao.cn{Path}", headers=headers, json=PostData
    )


"""POST方式匿名调用API"""


def PostWithoutTokenAPI(Path: str, PostData: dict) -> requests.Response:
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": UserAgent().random,
    }
    return requests.post(
        url=f"https://api.codemao.cn{Path}", headers=headers, json=PostData
    )


"""使用POST方式调用EDUAPI"""


def PostEduAPI(Path: str, PostData: dict, Token: str) -> requests.Response:
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": UserAgent().random,
        "authorization": f"Bearer {Token}",
    }
    return requests.post(
        url=f"https://eduzone.codemao.cn{Path}", headers=headers, json=PostData
    )


"""GET方式调用API"""


def GetAPI(Path: str, Token: str) -> requests.Response:
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": UserAgent().random,
        "authorization": Token,
    }
    return requests.get(url=f"https://api.codemao.cn{Path}", headers=headers)


"""GET方式匿名调用API"""


def GetWithoutTokenAPI(Path: str) -> requests.Response:
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": UserAgent().random,
    }
    return requests.get(url=f"https://api.codemao.cn{Path}", headers=headers)


"""PUT方式调用API"""


def PutAPI(Path: str, Token: str) -> requests.Response:
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": UserAgent().random,
        "authorization": Token,
    }
    return requests.put(url=f"https://api.codemao.cn{Path}", headers=headers)


"""确定Token数量"""


def CheckToken(Path: str) -> int:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return 0
    else:
        with open(Path, "r", encoding="utf-8") as f:
            FileLine = len(f.readlines())
            f.close()
            return FileLine


"""登录并获取用户Token"""


def GetUserToken(Username: str, Password: str) -> str | bool:
    response = PostWithoutTokenAPI(
        Path="/tiger/v3/web/accounts/login",
        PostData={"pid": "65edCTyg", "identity": Username, "password": Password},
    )
    if response.status_code == 200:
        return str(json.loads(response.text).get("auth", {}).get("token"))
    else:
        logging.error(
            f"请求失败，用户名：{Username}, 状态码: {response.status_code}, 响应: {response.text[:100]}"
        )
        return False


"""签订友好协议"""


def SignatureUser(Path: str) -> bool:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return False
    elif CheckToken(Path) == 0:
        logging.warning("可用的Token数为0")
        return False
    else:
        with open(Path, "r") as f:
            TokenList = [line.strip() for line in f if line.strip()]
            f.close()

        def CallToAPI_Sign(Token: str) -> bool:
            try:
                response = PostAPI(
                    Path="/nemo/v3/user/level/signature", PostData={}, Token=Token
                )

                if response.status_code == 200:
                    return True
                else:
                    logging.error(
                        f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                    )
                    return False
            except Exception as e:
                logging.error(f"请求异常: {str(e)}")
                return False

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(CallToAPI_Sign, TokenList))
            sum(results)

        return True


"""关注用户"""


def FollowUser(Path: str, UserID: str) -> bool:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return False
    elif CheckToken(Path) == 0:
        logging.warning("可用的Token数为0")
        return False
    else:
        with open(Path, "r") as f:
            TokenList = [line.strip() for line in f if line.strip()]
            f.close()

        def CallToAPI_Follow(Token: str) -> bool:
            try:
                response = PostAPI(
                    Path=f"/nemo/v2/user/{UserID}/follow", PostData={}, Token=Token
                )

                if response.status_code == 204:
                    return True
                else:
                    logging.error(
                        f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                    )
                    return False
            except Exception as e:
                logging.error(f"请求异常: {str(e)}")
                return False

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(CallToAPI_Follow, TokenList))
            sum(results)

        return True


"""获取用户所有的作品"""


def GetUserWork(UserID: str) -> str | bool:
    response = GetWithoutTokenAPI(
        f"/creation-tools/v2/user/center/work-list?type=newest&user_id={UserID}&offset=0&limit=1000"
    )
    if response.status_code == 200:
        ids = [str(item["id"]) for item in json.loads(response.text)["items"]]
        return " ".join(ids)
    else:
        logging.error(
            f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
        )
        return False


"""点赞作品"""


def LikeWork(Path: str, WorkID: str) -> bool:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return False
    elif CheckToken(Path) == 0:
        logging.warning("可用的Token数为0")
        return False
    else:
        with open(Path, "r") as f:
            TokenList = [line.strip() for line in f if line.strip()]
            f.close()

        def CallToAPI_Like(Token: str) -> bool:
            try:
                response = PostAPI(
                    Path=f"/nemo/v2/works/{WorkID}/like", PostData={}, Token=Token
                )

                if response.status_code == 200:
                    return True
                else:
                    logging.error(
                        f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                    )
                    return False
            except Exception as e:
                logging.error(f"请求异常: {str(e)}")
                return False

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(CallToAPI_Like, TokenList))
            sum(results)

        return True


"""收藏作品"""


def CollectionWork(Path: str, WorkID: str) -> bool:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return False
    elif CheckToken(Path) == 0:
        logging.warning("可用的Token数为0")
        return False
    else:
        with open(Path, "r") as f:
            TokenList = [line.strip() for line in f if line.strip()]
            f.close()

        def CallToAPI_Collection(Token: str) -> bool:
            try:
                response = PostAPI(
                    Path=f"/nemo/v2/works/{WorkID}/collection", PostData={}, Token=Token
                )

                if response.status_code == 200:
                    return True
                else:
                    logging.error(
                        f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                    )
                    return False
            except Exception as e:
                logging.error(f"请求异常: {str(e)}")
                return False

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(CallToAPI_Collection, TokenList))
            sum(results)

        return True


"""举报作品"""


def ReportWork(Path: str, WorkID: str, Reason: str, Describe: str) -> bool:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return False
    elif CheckToken(Path) == 0:
        logging.warning("可用的Token数为0")
        return False
    else:
        with open(Path, "r") as f:
            TokenList = []
            for i in range(report_readtoken_line):
                line = f.readline()
                if not line:
                    break
                line = line.strip()
                if line:
                    TokenList.append(line)
            f.close()

        def CallToAPI_ReportWork(Token: str) -> bool:
            try:
                response = PostAPI(
                    Path="/nemo/v2/report/work",
                    PostData={
                        "work_id": WorkID,
                        "report_reason": Reason,
                        "report_describe": Describe,
                    },
                    Token=Token,
                )

                if response.status_code == 200:
                    return True
                else:
                    logging.error(
                        f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                    )
                    return False
            except Exception as e:
                logging.error(f"请求异常: {str(e)}")
                return False

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(CallToAPI_ReportWork, TokenList))
            sum(results)

        return True


"""回复作品"""


def SendReviewToWork(Path: str, WorkID: str, ReviewText: str) -> bool:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return False
    elif CheckToken(Path) == 0:
        logging.warning("可用的Token数为0")
        return False
    else:
        with open(Path, "r") as f:
            TokenList = [line.strip() for line in f if line.strip()]
            f.close()

        def CallToAPI_Review(Token: str) -> bool:
            try:
                response = PostAPI(
                    Path=f"/creation-tools/v1/works/{WorkID}/comment",
                    PostData={"emoji_content": "", "content": ReviewText},
                    Token=Token,
                )

                if response.status_code == 201:
                    return True
                else:
                    logging.error(
                        f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                    )
                    return False
            except Exception as e:
                logging.error(f"请求异常: {str(e)}")
                return False

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(CallToAPI_Review, TokenList))
            sum(results)

        return True


"""置顶评论（越权）"""


def TopReview(Token: str, WorkID: str, CommentID: str) -> bool:
    try:
        response = PutAPI(
            f"/creation-tools/v1/works/{WorkID}/comment/{CommentID}/top", Token=Token
        )

        if response.status_code == 204:
            return True
        else:
            logging.error(
                f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
            )
            return False
    except Exception as e:
        logging.error(f"请求异常：{str(e)}")
        return False


"""浏览作品"""


def ViewWork(Token: str, WorkID: str) -> bool:
    try:
        response = GetAPI(f"/creation-tools/v1/works/{WorkID}", Token=Token)

        if response.status_code == 200:
            return True
        else:
            logging.error(
                f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
            )
            return False
    except Exception as e:
        logging.error(f"请求异常: {str(e)}")
        return False


"""再创作作品"""


def ForkWork(Path: str, WorkID: str) -> bool:
    if not os.path.exists(Path):
        logging.error(f"找不到Token文件: {Path}")
        return False
    elif CheckToken(Path) == 0:
        logging.warning("可用的Token数为0")
        return False
    else:
        with open(Path, "r") as f:
            TokenList = [line.strip() for line in f if line.strip()]
            f.close()

        def CallToAPI_Fork(Token: str) -> bool:
            try:
                response = PostAPI(
                    Path=f"/nemo/v2/works/{WorkID}/fork",
                    PostData={},
                    Token=Token,
                )

                if response.status_code == 200:
                    return True
                else:
                    logging.error(
                        f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                    )
                    return False
            except Exception as e:
                logging.error(f"请求异常: {str(e)}")
                return False

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(CallToAPI_Fork, TokenList))
            sum(results)

        return True


"""添加新的班级"""


def CreateClassOnEdu(Token: str, ClassName: str) -> str:
    try:
        response = PostEduAPI(
            Path="/edu/zone/class", PostData={"name": ClassName}, Token=Token
        )
        if response.status_code == 200:
            return str(json.loads(response.text).get("id"))
        else:
            logging.error(
                f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
            )
            return "0"
    except Exception as e:
        logging.error(f"请求异常: {str(e)}")
        return "0"


"""添加新的学生到班级"""


def CreateStudentOnEdu(
    Token: str, ClassID: str, StudentNameList: list[str]
) -> bytes | None | Any:
    try:
        response = PostEduAPI(
            Path=f"/edu/zone/class/{ClassID}/students",
            PostData={"student_names": StudentNameList},
            Token=Token,
        )
        if response.status_code == 200:
            return response.content
        else:
            logging.error(
                f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
            )
            return None
    except Exception as e:
        logging.error(f"请求异常: {str(e)}")
        return None


"""合并生成的表格"""


def MergeStudentXls(InputFolder: str, OutputFile: str) -> bool:
    try:
        if os.path.exists(InputFolder):
            main_wb = Workbook()
            main_ws = main_wb.active
            for filename in os.listdir(InputFolder):
                if filename.endswith(".xls"):
                    file_path = os.path.join(InputFolder, filename)
                    df = pd.read_excel(file_path, skiprows=3, dtype=str)
                    for index, row in df.iterrows():
                        main_ws.append(row.tolist())
            main_wb.save(OutputFile)
            return True
        else:
            logging.error(f"找不到输入的文件夹: {InputFolder}")
            return False
    except Exception as e:
        logging.error(f"出现异常: {str(e)}")
        return False


"""登录Edu账号"""


def LoginUseEdu(InputXlsx: str, OutputFile: str, Signature: bool) -> bool:
    CannotLogin = 0
    if Signature:
        logging.info("已开启同时签署用户协议功能！")
    if os.path.exists(InputXlsx):
        Sheet = load_workbook(InputXlsx).active
        AllAccount = Sheet.max_row
        logging.info(f"登录账号共{AllAccount}个")
        for _ in range(AllAccount):
            UserList = []
            PasswordList = []
            for Row in Sheet.iter_rows(
                min_row=1, min_col=3, max_col=4, values_only=True
            ):
                User, Password = Row
                UserList.append(User)
                PasswordList.append(Password)
            LoginReponse = GetUserToken(UserList[_], PasswordList[_])
            if not LoginReponse:
                CannotLogin += 1
            else:
                if Signature:
                    response = PostAPI(
                        Path="/nemo/v3/user/level/signature",
                        PostData={},
                        Token=LoginReponse,
                    )

                    if response.status_code != 200:
                        logging.error(
                            f"签署友好协议失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                        )
                with open(OutputFile, "a") as f:
                    f.write(LoginReponse + "\n")
        logging.warning(
            f"未成功登录数量: {CannotLogin}，占比{(CannotLogin / AllAccount) * 100}%"
        )
        return True
    else:
        logging.error(f"找不到输入的Xlsx文件: {InputXlsx}")
        return False


"""参数处理器"""


def CreateParser():
    parser = argparse.ArgumentParser(
        description=f"欢迎使用 CodemaoEDUTools! 当前版本: v{version}",
        epilog="示例: python3 main.py check-token",
    )

    global_group = parser.add_argument_group("全局参数")
    global_group.add_argument(
        "-tf", "--token-file", help="Token文件路径", default="tokens.txt"
    )

    subparsers = parser.add_subparsers(dest="command", help="可用命令")

    # Check_TokenFile(Path: str)
    check_tokenfile_parser = subparsers.add_parser(  # noqa: F841
        "check-token", help="查看一个Token文件内，有多少个Token（读取行数）"
    )

    # GetUserToken(Username: str, Password: str)
    getusertoken_parser = subparsers.add_parser(
        "get-token", help="登录以获取一个用户的Token"
    )

    getusertoken_parser.add_argument(
        "-u", "--username", required=True, help="用户名（手机号）"
    )

    getusertoken_parser.add_argument("-p", "--password", required=True, help="密码")

    # SignatureUser(Path: str)
    signature_parser = subparsers.add_parser(  # noqa: F841
        "signature",
        help="签订友好协议，推荐在使用其他功能前统一签订一次友好协议，防止出现无法请求的情况",
    )

    # FollowUser(Path: str, UserID: str)
    followuser_parser = subparsers.add_parser(
        "follow-user", help="批量关注一个用户，高情商就是刷粉丝"
    )

    followuser_parser.add_argument(
        "-uid", "--user-id", required=True, nargs="+", help="训练师编号"
    )

    # GetUserWork(UserID: str)
    getuserwork_parser = subparsers.add_parser("get-work", help="获取用户所有作品ID")

    getuserwork_parser.add_argument(
        "-uid", "--user-id", required=True, nargs="+", help="训练师编号"
    )

    # LikeWork(Path: str, WorkID: str)
    likework_parser = subparsers.add_parser("like-work", help="批量点赞一个作品")

    likework_parser.add_argument(
        "-wid", "--work-id", nargs="+", required=True, help="作品ID"
    )

    # CollectionWork(Path: str, WorkID: str)
    collectwork_parser = subparsers.add_parser("collect-work", help="批量收藏一个作品")

    collectwork_parser.add_argument(
        "-wid", "--work-id", nargs="+", required=True, help="作品ID"
    )

    # ReportWork(Path: str, WorkID: str, Reason: str, Describe: str)
    reportwork_parser = subparsers.add_parser(
        "report-work", help="批量举报一个作品，请勿大量Token举报"
    )

    reportwork_parser.add_argument(
        "-wid", "--work-id", nargs="+", required=True, help="作品ID"
    )

    reportwork_parser.add_argument(
        "-r", "--report-reason", required=True, help="原因，请参考文档给出的可用方式"
    )

    reportwork_parser.add_argument(
        "-d", "--report-describe", required=True, help="举报理由"
    )

    # SendReviewToWork(Path: str, WorkID: str, ReviewText: str)
    sendrevietowork_parser = subparsers.add_parser(
        "review-work", help="在一个作品下，批量发送同样的评论"
    )

    sendrevietowork_parser.add_argument(
        "-wid", "--work-id", required=True, nargs="+", help="作品ID"
    )

    sendrevietowork_parser.add_argument(
        "-r", "--review-text", required=True, nargs="+", help="评论内容"
    )

    # TopReview(Token: str, WorkID: str, CommentID: str)
    topreview_parser = subparsers.add_parser("review-top", help="越权置顶某个评论")

    topreview_parser.add_argument(
        "-t", "--one-token", required=True, help="一个可用Token"
    )

    topreview_parser.add_argument("-wid", "--work-id", required=True, help="作品ID")

    topreview_parser.add_argument("-cid", "--comment-id", required=True, help="评论ID")

    # ViewWork(Token: str, WorkID: str)
    viewwork_parser = subparsers.add_parser(
        "view-work",
        help="给作品加一个浏览，如果要一直刷，只需要循环这个函数就可以，一个Token就够",
    )

    viewwork_parser.add_argument(
        "-t", "--one-token", required=True, help="一个可用Token"
    )

    viewwork_parser.add_argument("-wid", "--work-id", required=True, help="作品ID")

    # ForkWork(Path: str, WorkID: str)
    forkwork_parser = subparsers.add_parser("fork-work", help="再创作一个作品")

    forkwork_parser.add_argument(
        "-wid", "--work-id", required=True, nargs="+", help="作品ID"
    )

    # CreateClassOnEdu(Token: str, ClassName: str)
    createclassedu_parser = subparsers.add_parser(
        "create-class", help="在Edu里添加一个新的班级"
    )

    createclassedu_parser.add_argument("-t", "--token", required=True, help="Edu Token")

    createclassedu_parser.add_argument(
        "-cn", "--class-name", required=True, help="班级名称，遵循官方命名规则"
    )

    # CreateStudentOnEdu(Token: str, ClassID: str, StudentNameList: list[str])
    createstudentedu_parser = subparsers.add_parser(
        "create-student", help="批量把创建新的学生并添加到班级内"
    )

    createstudentedu_parser.add_argument(
        "-t", "--token", required=True, help="Edu Token"
    )

    createstudentedu_parser.add_argument(
        "-cid", "--class-id", required=True, help="班级ID"
    )

    createstudentedu_parser.add_argument(
        "-sl",
        "--student-name-list",
        required=False,
        type=json.loads,
        default=student_names,
        help="学生名字的列表，最多100个学生，学生命名遵循官方命名规则",
    )

    createstudentedu_parser.add_argument(
        "-o",
        "--output-xls",
        required=False,
        default="output.xls",
        help="输出文件名，需要填写.xls后缀",
    )

    # MergeStudentXls(InputFolder: str, OutputFile:str)
    mergestudentxls_parser = subparsers.add_parser(
        "merge-xls", help="如果要合成为一个xlsx文件用于登录，请使用此函数"
    )

    mergestudentxls_parser.add_argument(
        "-if", "--input-xls-folder", required=True, help="含有多个.xls文件的文件夹"
    )

    mergestudentxls_parser.add_argument(
        "-o",
        "--output-xlsx",
        required=False,
        default="output_xls",
        help="输出文件名，需要填写.xlsx后缀",
    )

    # LoginUseEdu(InputXlsx:str, OutputFile:str)
    loginedu_parser = subparsers.add_parser(
        "login-edu",
        help="批量登录所有在xlsx内保存的账号密码，并打印Token到指定的文件内",
    )

    loginedu_parser.add_argument(
        "-i", "--input-xlsx", required=True, help="含有账号密码的xlsx表格文件的路径"
    )

    loginedu_parser.add_argument(
        "-s",
        "--signature-user",
        required=False,
        default=False,
        help="是否同时签署友好协议",
    )

    loginedu_parser.add_argument(
        "-o",
        "--output-txt",
        required=False,
        default="tokens.txt",
        help="输出文件名，需要填写.txt后缀",
    )

    # Version
    getversion_parser = subparsers.add_parser("version", help="获取CET版本")  # noqa: F841

    # End

    return parser


if __name__ == "__main__":
    logging.info("Welcome to CET")

    parser = CreateParser()
    args = parser.parse_args()

    # 主程序
    if args.command is None:
        logging.info('输入 "-h" 获得使用帮助')
        logging.info(
            "或者浏览文档：https://github.com/Wangs-official/CodemaoEDUTools/blob/main/doc/cli.md"
        )

    if args.command == "check-token":
        logging.info(f"可用Token数量: {CheckToken(args.token_file)}")

    if args.command == "get-token":
        logging.info(GetUserToken(args.username, args.password))

    if args.command == "signature":
        logging.info("请稍后...")
        if SignatureUser(args.token_file):
            logging.info("执行成功")

    if args.command == "follow-user":
        for i in args.user_id:
            logging.info(f"请稍后，正在执行：{i}")
            if FollowUser(args.token_file, i):
                logging.info("执行成功")

    if args.command == "get-work":
        for i in args.user_id:
            logging.info(f"请稍后，正在执行：{i}")
            if not GetUserWork(i):
                pass
            else:
                logging.info(f"用户：{i} 的作品列表：")
                logging.info(GetUserWork(i))

    if args.command == "like-work":
        for i in args.work_id:
            logging.info(f"请稍后，正在执行：{i}")
            if LikeWork(args.token_file, i):
                logging.info("执行成功")

    if args.command == "collect-work":
        for i in args.work_id:
            logging.info(f"请稍后，正在执行：{i}")
            if CollectionWork(args.token_file, i):
                logging.info("执行成功")

    if args.command == "report-work":
        for i in args.work_id:
            logging.info(f"请稍后，正在执行：{i}")
            if ReportWork(args.token_file, i, args.report_reason, args.report_describe):
                logging.info("执行成功")

    if args.command == "review-work":
        for i in args.work_id:
            for r in args.review_text:
                logging.info(f"请稍后，正在执行：{i} | 发送内容：{r}")
                if SendReviewToWork(args.token_file, i, r):
                    logging.info("执行成功")

    if args.command == "review-top":
        logging.info("请稍后...")
        if TopReview(args.one_token, args.work_id, args.comment_id):
            logging.info("执行成功")

    if args.command == "view-work":
        logging.info("请稍后...")
        if ViewWork(args.one_token, args.work_id):
            logging.info("执行成功")

    if args.command == "fork-work":
        for i in args.work_id:
            logging.info(f"请稍后，正在执行：{i}")
            if ForkWork(args.token_file, i):
                logging.info("执行成功")

    if args.command == "create-class":
        logging.info("请稍后...")
        logging.info(f"Class ID: {CreateClassOnEdu(args.token, args.class_name)}")

    if args.command == "create-student":
        logging.info("请稍后...")
        try:
            with open(args.output_xls, "wb") as f:
                f.write(
                    CreateStudentOnEdu(
                        args.token, args.class_id, args.student_name_list
                    )
                )
                f.close()
            logging.info(f"执行成功，学生密码表已保存到: {args.output_xls}")
        except TypeError:
            os.remove(args.output_xls)
            exit()

    if args.command == "merge-xls":
        logging.info("请稍后...")
        if MergeStudentXls(args.input_xls_folder, args.output_xlsx):
            logging.info(f"执行成功，合并的文件已保存到：{args.output_xlsx}")

    if args.command == "login-edu":
        logging.info("请稍后...")
        if LoginUseEdu(args.input_xlsx, args.output_txt, args.signature_user):
            logging.info(f"执行成功，已将登录的Token保存到：{args.output_txt}")

    if args.command == "version":
        logging.info(
            f"CET版本: v{version}\nhttps://github.com/Wangs-official/CodemaoEDUTools/"
        )
